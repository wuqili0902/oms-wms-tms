"""Barcode API router.

/templates routes MUST be before /{gtin} to avoid FastAPI path-parameter matching.
"""
import io
import uuid

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from src.barcode import service as barcode_service
from src.barcode.excel_barcode import (
    export_csv_barcodes,
    generate_barcode_zip,
)
from src.barcode.schemas import (
    BarcodeGenerateRequest,
    BarcodeResponse,
    BarcodeScanRequest,
    BarcodeValidateRequest,
    LabelTemplateCreate,
    LabelTemplateResponse,
)
from src.core.database import get_db
from src.core.dependencies import get_current_user
from src.core.exceptions import ValidationException

router = APIRouter(prefix="/barcode", tags=["barcode"])


# ── Core barcode endpoints ─────────────────────────────────────────────


@router.post("/generate", response_model=BarcodeResponse, status_code=status.HTTP_201_CREATED)
async def generate_barcode(
    data: BarcodeGenerateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    try:
        rec = await barcode_service.generate_barcode(db, data.model_dump())
        return {k: rec[k] for k in ["id", "gtin", "entity_type", "entity_id", "format", "raw_data", "created_at"]}
    except ValidationException as e:
        raise HTTPException(status_code=422, detail=str(e))


@router.post("/validate")
async def validate_barcode(data: BarcodeValidateRequest, current_user: dict = Depends(get_current_user)):
    return barcode_service.validate_gtin(data.gtin)


@router.post("/scan", response_model=BarcodeResponse, status_code=status.HTTP_201_CREATED)
async def scan_barcode(
    data: BarcodeScanRequest,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    rec = await barcode_service.record_scan(db, data.model_dump())
    return {k: rec[k] for k in ["id", "gtin", "entity_type", "entity_id", "format", "raw_data", "created_at"]}


# ── Excel 条码批量生成 ───────────────────────────────────────────────


@router.post("/excel/upload", response_model=dict[str, str | int])
async def upload_barcode_excel(
    file: UploadFile,
    current_user: dict = Depends(get_current_user),
):
    """上传 .xlsx → 返回包含所有条码图片的 ZIP"""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="仅支持 .xlsx 文件")

    try:
        zip_bytes = generate_barcode_zip(file)
        filename = f"barcodes_{uuid.uuid4().hex[:8]}.zip"
        return {"filename": filename, "size": len(zip_bytes)}
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))



    """上传 .xlsx → 下载 CSV（含 SKU + GTIN）"""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="仅支持 .xlsx 文件")

    try:
        wb = load_workbook(io.BytesIO(file.file.read()))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=422, detail="空工作表")

        headers = [str(h).strip().lower() for h in rows[0]]
        data_rows = []
        import pandas as pd  # noqa: F811

        for row in rows[1:]:
            mapping = {col: (row[i] if i < len(row) else None) for i, col in enumerate(headers)}
            data_rows.append(mapping)

        df = pd.DataFrame(data_rows)
        csv_str = export_csv_barcodes(df)
        return csv_str
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


# ── ZPL 热敏打印标签 ───────────────────────────────────────────────



    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="仅支持 .xlsx 文件")

    try:
        from src.barcode.excel_barcode import export_zpl_labels as _zpl_export

        wb = load_workbook(io.BytesIO(file.file.read()))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=422, detail="空工作表")

        headers = [str(h).strip().lower() for h in rows[0]]
        data_rows = []
        import pandas as pd  # noqa: F811

        for row in rows[1:]:
            mapping = {col: (row[i] if i < len(row) else None) for i, col in enumerate(headers)}
            data_rows.append(mapping)

        df = pd.DataFrame(data_rows)
        zpl_str = _zpl_export(df)
        return zpl_str
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


# ── Download generated ZIP file (called after excel/upload) ──────────────


@router.get("/download/<filename>")
async def download_barcode_file(
    filename: str,
    current_user: dict = Depends(get_current_user),
):
    """下载之前通过 /excel/upload 生成的 ZIP 文件"""
    filepath = f"/tmp/barcodes/{filename}"
    try:
        with open(filepath, "rb") as f:
            data = f.read()
        from fastapi.responses import StreamingResponse

        return StreamingResponse(io.BytesIO(data), media_type="application/zip", headers={"Content-Disposition": f"attachment; filename={filename}"})
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"File not found: {filename}")


# ── Templates — MUST be before /{gtin} ═══════════════════════════════


@router.post("/templates", response_model=LabelTemplateResponse, status_code=status.HTTP_201_CREATED)
async def create_template(
    data: LabelTemplateCreate,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    try:
        tpl = await barcode_service.create_template(db, data.model_dump())
        return LabelTemplateResponse(**tpl)
    except ValidationException as e:
        raise HTTPException(status_code=422, detail=str(e))


@router.get("/templates", response_model=list[LabelTemplateResponse])
async def list_templates(
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    templates = await barcode_service.list_templates(db)
    return [LabelTemplateResponse(**t) for t in templates]


# ── GTIN lookup — must be last ═══════════════════════════════════════


@router.get("/{gtin}", response_model=list[BarcodeResponse])
async def get_by_gtin(
    gtin: str,
    db: AsyncSession = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    records = await barcode_service.get_by_gtin(db, gtin)
    return [
        {k: r[k] for k in ["id", "gtin", "entity_type", "entity_id", "format", "raw_data", "created_at"]}
        for r in records
    ]
