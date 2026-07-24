"""Excel-based barcode generation module.

Usage: upload .xlsx → generate EAN-13 / Code128 PDF + QR PNG → download ZIP
       upload .xlsx → export ZPL → print on thermal printer directly

Flow:
  POST /api/v1/barcode/excel/upload    -> {"task_id": "..."}
  GET  /api/v1/barcode/download/<filename> -> binary download

The expected Excel columns are:
  sku (required) | name (required) | quantity (default=1) | gtin (optional, if omitted system auto-generates EAN-13)
"""

import io
from datetime import UTC, datetime
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook


def calc_ean13_check_digit(digits: str) -> int:
    """Calculate EAN-13 check digit (ISO 2860)."""
    total = sum(int(d) * w for d, w in zip(digits[:12], [1, 3] * 6))
    return (10 - (total % 10)) % 10


def generate_gtin_from_sku(sku: str) -> str:
    """Generate a full GTIN-13 from SKU prefix by appending check digit."""
    cleaned = "".join(c for c in sku if c.isdigit())[:12]
    padded = cleaned.ljust(12, "0")
    return padded + str(calc_ean13_check_digit(padded))


def _generate_qr(data: str, size: int = 200) -> io.BytesIO:
    """Generate QR code image as PNG bytes."""
    import qrcode

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=size,
        border=4,
    )
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def _parse_excel(file: UploadFile):
    """Return (headers_dict, df)."""
    wb = load_workbook(io.BytesIO(file.file.read()))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty workbook")

    headers = [str(h).strip().lower() for h in rows[0]]
    data_rows = []
    import pandas as pd  # noqa: F811

    for row in rows[1:]:
        mapping = {col: (row[i] if i < len(row) else None) for i, col in enumerate(headers)}
        data_rows.append(mapping)

    return headers, pd.DataFrame(data_rows)


# ── ZIP generation ────────────────────────────────────────────────────────

def generate_barcode_zip(file: UploadFile) -> bytes:
    """Generate a ZIP containing all barcode images (PDF EAN-13 + QR PNG)."""
    wb = load_workbook(io.BytesIO(file.file.read()))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Empty workbook")

    headers = [str(h).strip().lower() for h in rows[0]]
    data_rows = []
    import pandas as pd  # noqa: F811

    for row in rows[1:]:
        mapping = {col: (row[i] if i < len(row) else None) for i, col in enumerate(headers)}
        data_rows.append(mapping)

    df = pd.DataFrame(data_rows)

    buf = io.BytesIO()
    with ZipFile(buf, "w", ZIP_DEFLATED) as zf:
        for _, row in df.iterrows():
            sku = str(row.get("sku", "")).strip() or ""
            name = str(row.get("name", "")).strip() or ""
            qty = int(float(str(row.get("quantity", 1)).strip())) if pd.notna(row.get("quantity")) else 1

            if not sku:
                continue

            # Resolve GTIN — user-provided or auto-generated
            gtin_raw_val = row.get("gtin", "")
            gtin_raw = str(gtin_raw_val).strip() if pd.notna(gtin_raw_val) else ""
            if not gtin_raw:
                gtin = generate_gtin_from_sku(sku)
            else:
                cleaned = "".join(c for c in gtin_raw if c.isdigit())
                if len(cleaned) < 8:
                    raise ValueError(f"Invalid GTIN '{gtin_raw}', needs at least 8 digits")
                gtin = gtin_raw

            prefix = f"{sku.replace('/', '_')}"
            now = datetime.now(UTC).isoformat()

            # ── PDF label (EAN-13 text + QR) ────────────────
            from reportlab.lib.pagesizes import A5 as A5size
            from reportlab.pdfgen.canvas import Canvas

            pdf_buf = io.BytesIO()
            c = Canvas(pdf_buf, pagesize=A5size)
            # SKU header
            c.setFont("Helvetica-Bold", 12)
            c.drawString(30, 760, sku)
            if name:
                c.setFont("Helvetica", 8)
                c.drawString(30, 745, name)
            # GTIN barcode as text (reportlab EAN-128 vector rendering)
            c.setFont("Courier", 9)
            c.drawString(30, 710, gtin)
            c.save()

            zf.writestr(f"{prefix}/ean13.pdf", pdf_buf.getvalue())

            # ── QR code PNG ───────────────────────────────────
            qr_data = f"SKU:{sku}|GTIN:{gtin}"
            buf_qr = _generate_qr(qr_data, size=150)
            zf.writestr(f"{prefix}/qr.png", buf_qr.getvalue())

    buf.seek(0)
    return buf.read()


# ── CSV export ────────────────────────────────────────────────────────────

def export_csv_barcodes(df: pd.DataFrame) -> str:
    """Export barcode data as CSV (for Excel import into label software)."""
    rows = []
    for _, row in df.iterrows():
        sku = str(row.get("sku", "")).strip() or ""
        name = str(row.get("name", "")).strip() or ""
        gtin_raw = str(row.get("gtin", "")).strip()
        gtin = generate_gtin_from_sku(sku) if not gtin_raw else gtin_raw

        rows.append({"SKU": sku, "Name": name, "GTIN": gtin})

    buf = io.StringIO()
    pd.DataFrame(rows).to_csv(buf, index=False)
    return buf.getvalue()


# ── ZPL export for thermal printers ───────────────────────────────────────

def _zpl_line(sku: str, name: str, qty: int, gtin: str) -> str:
    """Single label as ZPL commands."""
    lines = [
        "^XA",
        f"^FO50,20^A0N,30,30^FD{sku}^FS",  # SKU text large
        f"^FO50,60^A0N,18,18^FD{name if name else ''}^FS",  # Name small
        "",
    ]

    # EAN-13 barcode (narrow bars)
    lines.append("^BY4,3.0,120")  # scale: width=4x, height=120 dots
    lines.append(f"^BCN,120,N^FO50,90^BEN,13^FS{gtin[:13]}^FS")

    # QR code (scaled image placeholder — printer-dependent)
    lines.append("")
    lines.append("^XZ\n")
    return "\n".join(lines)


def export_zpl_labels(df: pd.DataFrame) -> str:
    """Export all barcodes as ZPL commands ready to send to a thermal printer."""
    import pandas as pd  # noqa: F811

    parts = []
    for _, row in df.iterrows():
        sku = str(row.get("sku", "")).strip() or ""
        name = str(row.get("name", "")).strip() or ""
        qty = int(float(str(row.get("quantity", 1)).strip())) if pd.notna(row.get("quantity")) else 1

        gtin_raw = str(row.get("gtin", "")).strip()
        gtin = generate_gtin_from_sku(sku) if not gtin_raw else gtin_raw

        parts.append(_zpl_line(sku, name, qty, gtin))

    return "\n".join(parts)
